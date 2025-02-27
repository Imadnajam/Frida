import os
import io
import re
import json
import base64
from typing import Dict, List, Optional, Tuple, Union

from fastapi import APIRouter, File, UploadFile, HTTPException, Form
from fastapi.responses import JSONResponse
from pydantic import BaseModel

# File processing libraries
import docx2txt
import csv
import PyPDF2
import mammoth
import openpyxl
import markdown
import html2text
import xml.etree.ElementTree as ET
from bs4 import BeautifulSoup
from PIL import Image


# Type definitions for clarity
class ConversionResult(BaseModel):
    markdown: str
    metadata: Dict[str, Union[str, int, float, List[str], Dict[str, str]]]
    content_type: str
    preview: Optional[str] = None


# Constants
MAX_FILE_SIZE = 20 * 1024 * 1024
SUPPORTED_FORMATS = {
    # Documents
    "application/pdf": "pdf",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document": "docx",
    "application/msword": "doc",
    "text/plain": "txt",
    "text/markdown": "md",
    "text/html": "html",
    "text/xml": "xml",
    "application/json": "json",
    "text/csv": "csv",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": "xlsx",
    # Images
    "image/jpeg": "jpg",
    "image/png": "png",
    "image/gif": "gif",
    "image/svg+xml": "svg",
    # Other
    "application/rtf": "rtf",
}

# File converter router
router = APIRouter()


def create_error_response(status_code: int, detail: str) -> JSONResponse:
    """Create a standardized error response"""
    return JSONResponse(
        status_code=status_code, content={"success": False, "error": detail}
    )


def clean_text(text: str) -> str:
    """Clean extracted text for better markdown compatibility"""
    # Remove excessive whitespace
    cleaned_text = re.sub(r"\s+", " ", text)
    # Remove control characters
    cleaned_text = re.sub(r"[\x00-\x1F\x7F]", "", cleaned_text)
    # Replace tabs with spaces
    cleaned_text = cleaned_text.replace("\t", "    ")
    # Ensure proper line breaks
    cleaned_text = re.sub(r"(\r\n|\r|\n){3,}", "\n\n", cleaned_text)
    return cleaned_text.strip()


def extract_metadata(file: UploadFile, content_type: str) -> Dict:
    """Extract metadata from files based on their type"""
    basic_metadata = {
        "filename": file.filename,
        "content_type": content_type,
        "size": 0,  # Will be updated after processing
    }

    # Return for now, will be extended when processing specific file types
    return basic_metadata


async def convert_pdf_to_markdown(file: UploadFile) -> ConversionResult:
    """Convert PDF files to markdown"""
    try:
        pdf_reader = PyPDF2.PdfReader(file.file)
        text = ""
        metadata = {
            "page_count": len(pdf_reader.pages),
            "title": file.filename,
        }

        # Extract text from all pages
        for page_num, page in enumerate(pdf_reader.pages):
            page_text = page.extract_text()
            if page_text:
                text += f"\n\n## Page {page_num + 1}\n\n{page_text}"

        # Get document info if available
        if pdf_reader.metadata:
            info = pdf_reader.metadata
            if info.title:
                metadata["title"] = info.title
            if info.author:
                metadata["author"] = info.author
            if info.subject:
                metadata["subject"] = info.subject
            if info.creator:
                metadata["creator"] = info.creator

        # Clean the text
        markdown_text = clean_text(text)

        # Format with proper markdown
        markdown_text = f"# {metadata.get('title', 'PDF Document')}\n\n{markdown_text}"

        return ConversionResult(
            markdown=markdown_text, metadata=metadata, content_type="application/pdf"
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error converting PDF: {str(e)}")


async def convert_docx_to_markdown(file: UploadFile) -> ConversionResult:
    """Convert DOCX files to markdown"""
    try:
        # Save temporarily
        content = await file.read()
        file.file.seek(0)  # Reset file pointer

        # Use mammoth for better conversion with styles
        result = mammoth.convert_to_markdown(io.BytesIO(content))
        markdown_text = result.value

        # Extract basic metadata
        metadata = {
            "filename": file.filename,
            "content_type": file.content_type,
            "size": len(content),
            "messages": [message.message for message in result.messages],
        }

        return ConversionResult(
            markdown=markdown_text,
            metadata=metadata,
            content_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error converting DOCX: {str(e)}")


async def convert_txt_to_markdown(file: UploadFile) -> ConversionResult:
    """Convert plain text files to markdown"""
    try:
        content = await file.read()
        file.file.seek(0)

        # Decode text content
        text = content.decode("utf-8", errors="replace")

        # For text files, we'll add basic formatting
        lines = text.split("\n")

        # Try to identify a title (first non-empty line)
        title = file.filename
        for line in lines:
            if line.strip():
                title = line.strip()
                break

        markdown_text = f"# {title}\n\n{text}"

        return ConversionResult(
            markdown=markdown_text,
            metadata={
                "filename": file.filename,
                "content_type": "text/plain",
                "size": len(content),
            },
            content_type="text/plain",
        )
    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Error converting text file: {str(e)}"
        )


async def convert_html_to_markdown(file: UploadFile) -> ConversionResult:
    """Convert HTML files to markdown"""
    try:
        content = await file.read()
        file.file.seek(0)

        # Decode HTML content
        html_content = content.decode("utf-8", errors="replace")

        # Use html2text for conversion
        converter = html2text.HTML2Text()
        converter.ignore_links = False
        converter.ignore_images = False
        converter.body_width = 0  # No wrapping

        markdown_text = converter.handle(html_content)

        # Extract title if available
        title = file.filename
        soup = BeautifulSoup(html_content, "html.parser")
        if soup.title and soup.title.string:
            title = soup.title.string.strip()

        metadata = {
            "filename": file.filename,
            "content_type": "text/html",
            "size": len(content),
            "title": title,
        }

        return ConversionResult(
            markdown=markdown_text, metadata=metadata, content_type="text/html"
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error converting HTML: {str(e)}")


async def convert_csv_to_markdown(file: UploadFile) -> ConversionResult:
    """Convert CSV files to markdown tables"""
    try:
        content = await file.read()
        file.file.seek(0)

        # Decode CSV content
        text = content.decode("utf-8", errors="replace")

        # Parse CSV
        reader = csv.reader(io.StringIO(text))
        rows = list(reader)

        if not rows:
            return ConversionResult(
                markdown="*Empty CSV file*",
                metadata={
                    "filename": file.filename,
                    "content_type": "text/csv",
                    "size": len(content),
                },
                content_type="text/csv",
            )

        # Create markdown table
        header = rows[0]
        markdown_table = "| " + " | ".join(header) + " |\n"
        markdown_table += "| " + " | ".join(["---" for _ in header]) + " |\n"

        # Add data rows
        for row in rows[1:]:
            # Ensure row has the same number of columns as header
            padded_row = row + [""] * (len(header) - len(row))
            markdown_table += "| " + " | ".join(padded_row[: len(header)]) + " |\n"

        # Add title
        markdown_text = f"# {file.filename}\n\n{markdown_table}"

        return ConversionResult(
            markdown=markdown_text,
            metadata={
                "filename": file.filename,
                "content_type": "text/csv",
                "size": len(content),
                "rows": len(rows),
                "columns": len(header),
            },
            content_type="text/csv",
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error converting CSV: {str(e)}")


async def convert_json_to_markdown(file: UploadFile) -> ConversionResult:
    """Convert JSON files to markdown with code blocks"""
    try:
        content = await file.read()
        file.file.seek(0)

        # Decode JSON content
        text = content.decode("utf-8", errors="replace")

        # Parse JSON to ensure validity and for pretty formatting
        json_data = json.loads(text)
        pretty_json = json.dumps(json_data, indent=2)

        # Create markdown with code block
        markdown_text = f"# {file.filename}\n\n```json\n{pretty_json}\n```"

        # Extract some basic metadata
        metadata = {
            "filename": file.filename,
            "content_type": "application/json",
            "size": len(content),
        }

        # If JSON is an object, add top-level keys to metadata
        if isinstance(json_data, dict):
            metadata["keys"] = list(json_data.keys())
        # If JSON is an array, add length
        elif isinstance(json_data, list):
            metadata["items"] = len(json_data)

        return ConversionResult(
            markdown=markdown_text, metadata=metadata, content_type="application/json"
        )
    except json.JSONDecodeError as e:
        # If JSON is invalid, return the error in markdown
        markdown_text = f"# {file.filename}\n\n**Error parsing JSON:**\n\n{str(e)}\n\n```\n{text[:1000]}...\n```"
        return ConversionResult(
            markdown=markdown_text,
            metadata={
                "filename": file.filename,
                "content_type": "application/json",
                "size": len(content),
                "error": str(e),
            },
            content_type="application/json",
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error converting JSON: {str(e)}")


async def convert_xlsx_to_markdown(file: UploadFile) -> ConversionResult:
    """Convert Excel files to markdown tables"""
    try:
        content = await file.read()
        file.file.seek(0)

        # Open the workbook
        workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True)

        markdown_text = f"# {file.filename}\n\n"

        # Process each worksheet
        sheets_data = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet_md = f"## Sheet: {sheet_name}\n\n"

            # Get the data from the sheet
            data = []
            for row in sheet.iter_rows(values_only=True):
                if any(cell is not None for cell in row):  # Skip completely empty rows
                    data.append([str(cell) if cell is not None else "" for cell in row])

            if not data:
                sheet_md += "*Empty sheet*\n\n"
                sheets_data.append({"name": sheet_name, "rows": 0, "columns": 0})
                continue

            # Create markdown table
            max_cols = max(len(row) for row in data)
            sheet_md += (
                "| " + " | ".join([f"Column {i+1}" for i in range(max_cols)]) + " |\n"
            )
            sheet_md += "| " + " | ".join(["---" for _ in range(max_cols)]) + " |\n"

            # Add data rows (limit to first 100 rows for large sheets)
            row_limit = min(100, len(data))
            for row in data[:row_limit]:
                # Ensure row has the right number of columns
                padded_row = row + [""] * (max_cols - len(row))
                sheet_md += "| " + " | ".join(padded_row) + " |\n"

            if len(data) > row_limit:
                sheet_md += "\n*Note: Only showing first 100 rows*\n\n"

            markdown_text += sheet_md + "\n\n"
            sheets_data.append(
                {"name": sheet_name, "rows": len(data), "columns": max_cols}
            )

        return ConversionResult(
            markdown=markdown_text,
            metadata={
                "filename": file.filename,
                "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "size": len(content),
                "sheets": sheets_data,
            },
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Error converting Excel file: {str(e)}"
        )


async def convert_image_to_markdown(file: UploadFile) -> ConversionResult:
    """Convert image files to markdown with embedded Base64 (for small images) or description"""
    try:
        content = await file.read()
        file.file.seek(0)

        # For images, if small enough, we'll convert to base64 for embedding
        # Otherwise, we'll just describe the image
        if len(content) < 1024 * 1024:  # Less than 1MB
            try:
                img = Image.open(io.BytesIO(content))
                width, height = img.size
                format_lower = img.format.lower() if img.format else "unknown"

                # Convert to base64 for preview
                img_base64 = base64.b64encode(content).decode("utf-8")
                image_md = (
                    f"![{file.filename}](data:image/{format_lower};base64,{img_base64})"
                )

                markdown_text = f"# Image: {file.filename}\n\n{image_md}\n\n**Details:**\n\n- Width: {width}px\n- Height: {height}px\n- Format: {img.format}\n- Mode: {img.mode}"

                return ConversionResult(
                    markdown=markdown_text,
                    metadata={
                        "filename": file.filename,
                        "content_type": file.content_type,
                        "size": len(content),
                        "width": width,
                        "height": height,
                        "format": img.format,
                    },
                    content_type=file.content_type,
                    preview=f"data:image/{format_lower};base64,{img_base64}",
                )
            except Exception as img_error:
                # If image processing fails, provide basic info
                markdown_text = f"# Image: {file.filename}\n\n*Could not process image for preview: {str(img_error)}*\n\n**Details:**\n\n- Size: {len(content)} bytes"
        else:
            markdown_text = f"# Image: {file.filename}\n\n*Image too large for preview (size: {len(content)/1024/1024:.2f} MB)*"

        return ConversionResult(
            markdown=markdown_text,
            metadata={
                "filename": file.filename,
                "content_type": file.content_type,
                "size": len(content),
            },
            content_type=file.content_type,
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error converting image: {str(e)}")


async def convert_xml_to_markdown(file: UploadFile) -> ConversionResult:
    """Convert XML files to markdown with syntax highlighting"""
    try:
        content = await file.read()
        file.file.seek(0)

        # Decode XML content
        xml_text = content.decode("utf-8", errors="replace")

        # Try to pretty-print the XML
        try:
            root = ET.fromstring(xml_text)
            # Use minidom for pretty printing
            from xml.dom import minidom

            pretty_xml = minidom.parseString(ET.tostring(root)).toprettyxml()
        except:
            pretty_xml = xml_text  # Use original if parsing fails

        # Create markdown with code block
        markdown_text = f"# {file.filename}\n\n```xml\n{pretty_xml}\n```"

        return ConversionResult(
            markdown=markdown_text,
            metadata={
                "filename": file.filename,
                "content_type": "text/xml",
                "size": len(content),
            },
            content_type="text/xml",
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error converting XML: {str(e)}")


async def convert_unknown_to_markdown(file: UploadFile) -> ConversionResult:
    """Handle unknown file types by providing basic info"""
    try:
        content = await file.read()
        file.file.seek(0)

        # Try to detect if it's a text file
        is_text = True
        try:
            text = content.decode("utf-8", errors="strict")
            # If more than 10% of characters are null or control, likely binary
            control_chars = sum(
                1 for c in text if ord(c) < 32 and c not in ["\n", "\r", "\t"]
            )
            if control_chars / len(text) > 0.1:
                is_text = False
        except UnicodeDecodeError:
            is_text = False

        if is_text:
            # It's a text file we can display
            text = content.decode("utf-8", errors="replace")
            markdown_text = f"# File: {file.filename}\n\n```\n{text[:10000]}\n```"
            if len(text) > 10000:
                markdown_text += (
                    "\n\n*Note: File truncated, showing first 10,000 characters*"
                )
        else:
            # It's a binary file we can't display
            file_size = len(content)
            markdown_text = f"# Binary File: {file.filename}\n\n- **Size**: {file_size} bytes ({file_size/1024/1024:.2f} MB)\n- **Type**: {file.content_type or 'Unknown'}\n\n*This file appears to be binary and cannot be displayed as text.*"

        return ConversionResult(
            markdown=markdown_text,
            metadata={
                "filename": file.filename,
                "content_type": file.content_type or "application/octet-stream",
                "size": len(content),
            },
            content_type=file.content_type or "application/octet-stream",
        )
    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Error processing unknown file: {str(e)}"
        )


@router.post("/convert")
async def convert_to_markdown(file: UploadFile = File(...)):
    """
    Convert various file formats to Markdown.
    Supports PDF, DOCX, TXT, HTML, CSV, JSON, XLSX, images, and more.
    """
    try:
        # Validate file presence
        if not file:
            return create_error_response(400, "No file uploaded")

        # Get file content type
        content_type = file.content_type or "application/octet-stream"

        # Validate file size
        file.file.seek(0, 2)  # Move to the end of the file
        file_size = file.file.tell()  # Get file size
        file.file.seek(0)  # Reset file pointer
        if file_size > MAX_FILE_SIZE:
            return create_error_response(
                400, f"File size exceeds {MAX_FILE_SIZE / 1024 / 1024} MB"
            )

        # Choose converter based on content type
        if content_type == "application/pdf":
            result = await convert_pdf_to_markdown(file)
        elif content_type in [
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            "application/msword",
        ]:
            result = await convert_docx_to_markdown(file)
        elif content_type == "text/plain":
            result = await convert_txt_to_markdown(file)
        elif content_type == "text/html":
            result = await convert_html_to_markdown(file)
        elif content_type == "text/csv":
            result = await convert_csv_to_markdown(file)
        elif content_type == "application/json":
            result = await convert_json_to_markdown(file)
        elif (
            content_type
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ):
            result = await convert_xlsx_to_markdown(file)
        elif content_type in ["image/jpeg", "image/png", "image/gif", "image/svg+xml"]:
            result = await convert_image_to_markdown(file)
        elif content_type == "text/xml":
            result = await convert_xml_to_markdown(file)
        else:
            # Try to handle unknown formats
            result = await convert_unknown_to_markdown(file)

        # Format the markdown as a code block for the API response
        markdown_code_block = f"```markdown\n{result.markdown}\n```"

        # Return the response
        return JSONResponse(
            status_code=200,
            content={
                "success": True,
                "message": "File converted successfully",
                "markdownContent": markdown_code_block,
                "rawMarkdown": result.markdown,
                "metadata": result.metadata,
                "preview": result.preview,
                "sourceFormat": SUPPORTED_FORMATS.get(content_type, "unknown"),
            },
        )

    except HTTPException as he:
        # Re-raise HTTP exceptions
        raise he
    except Exception as e:
        print(f"Unexpected error during conversion: {str(e)}")
        return create_error_response(500, f"Failed to convert file: {str(e)}")


@router.get("/supported-formats")
async def get_supported_formats():
    """Get list of supported file formats for conversion"""
    formats = []
    for mime, extension in SUPPORTED_FORMATS.items():
        formats.append(
            {
                "mime": mime,
                "extension": extension,
                "description": f"{extension.upper()} files",
            }
        )

    return JSONResponse(
        status_code=200,
        content={
            "success": True,
            "formats": formats,
            "maxFileSize": MAX_FILE_SIZE,
            "maxFileSizeMB": MAX_FILE_SIZE / 1024 / 1024,
        },
    )
